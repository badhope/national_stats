#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报告导出模块
宏观经济数据分析平台 - PDF/Excel 报告生成
"""

import os
import json
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportExporter:
    """报告导出器基类"""
    
    def __init__(self, output_dir: str = "output/reports"):
        self.output_dir = output_dir
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """确保输出目录存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def _get_timestamp(self) -> str:
        """获取时间戳"""
        return datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def export(self, data: Dict[str, Any], format: str = 'excel') -> str:
        """导出报告
        
        Args:
            data: 要导出的数据
            format: 导出格式 ('excel', 'pdf', 'csv', 'json')
        
        Returns:
            生成的文件路径
        """
        if format == 'excel':
            return self._export_excel(data)
        elif format == 'pdf':
            return self._export_pdf(data)
        elif format == 'csv':
            return self._export_csv(data)
        elif format == 'json':
            return self._export_json(data)
        else:
            raise ValueError(f"不支持的格式: {format}")


class ExcelExporter(ReportExporter):
    """Excel 报告导出器"""
    
    def _export_excel(self, data: Dict[str, Any]) -> str:
        """导出为 Excel 文件"""
        if not HAS_OPENPYXL:
            return self._export_csv(data)
        
        filename = f"economic_report_{self._get_timestamp()}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = openpyxl.Workbook()
        
        ws_summary = wb.active
        ws_summary.title = "数据摘要"
        self._write_summary_sheet(ws_summary, data)
        
        if 'gdp' in data:
            ws_gdp = wb.create_sheet("GDP数据")
            self._write_data_sheet(ws_gdp, data['gdp'], "国内生产总值")
        
        if 'cpi' in data:
            ws_cpi = wb.create_sheet("CPI数据")
            self._write_data_sheet(ws_cpi, data['cpi'], "居民消费价格指数")
        
        if 'employment' in data:
            ws_emp = wb.create_sheet("就业数据")
            self._write_data_sheet(ws_emp, data['employment'], "城镇就业情况")
        
        wb.save(filepath)
        return filepath
    
    def _write_summary_sheet(self, ws, data: Dict):
        """写入摘要工作表"""
        ws['A1'] = "宏观经济数据分析报告"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "报告生成时间"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "数据指标"
        ws['B4'] = "最新值"
        ws['C4'] = "同比变化"
        ws['D4'] = "趋势"
        
        row = 5
        for key, value in data.items():
            if isinstance(value, dict):
                ws[f'A{row}'] = value.get('title', key.upper())
                ws[f'B{row}'] = value.get('latest', 'N/A')
                ws[f'C{row}'] = value.get('change', 'N/A')
                ws[f'D{row}'] = value.get('trend', 'N/A')
                row += 1
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _write_data_sheet(self, ws, data: Dict, title: str):
        """写入数据工作表"""
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['年份', '数值', '同比增长率(%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        years = data.get('years', [])
        values = data.get('values', [])
        changes = data.get('changes', [])
        
        for i, (year, value, change) in enumerate(zip(years, values, changes), 4):
            ws.cell(row=i, column=1, value=year)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=3, value=change)
        
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 18


class PDFExporter(ReportExporter):
    """PDF 报告导出器"""
    
    def _export_pdf(self, data: Dict[str, Any]) -> str:
        """导出为 PDF 文件"""
        if not HAS_REPORTLAB:
            raise ImportError("请安装 reportlab: pip install reportlab")
        
        filename = f"economic_report_{self._get_timestamp()}.pdf"
        filepath = os.path.join(self.output_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30
        )
        
        story.append(Paragraph("宏观经济数据分析报告", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        story.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        for key, value in data.items():
            if isinstance(value, dict):
                story.append(Paragraph(value.get('title', key.upper()), styles['Heading2']))
                story.append(Spacer(1, 0.1*inch))
                
                table_data = [['指标', '数值']]
                if 'latest' in value:
                    table_data.append(['最新值', str(value['latest'])])
                if 'change' in value:
                    table_data.append(['同比变化', str(value['change'])])
                if 'trend' in value:
                    table_data.append(['趋势', str(value['trend'])])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
                story.append(Spacer(1, 0.3*inch))
        
        doc.build(story)
        return filepath
    
    def _export_csv(self, data: Dict[str, Any]) -> str:
        """导出为 CSV 文件"""
        return self._export_csv_format(data)
    
    def _export_json(self, data: Dict[str, Any]) -> str:
        """导出为 JSON 文件"""
        filename = f"economic_report_{self._get_timestamp()}.json"
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump({
                'generated_at': datetime.now().isoformat(),
                'data': data
            }, f, ensure_ascii=False, indent=2)
        
        return filepath


class CSVExporter(ReportExporter):
    """CSV 导出器"""
    
    def _export_csv_format(self, data: Dict[str, Any]) -> str:
        """导出为 CSV 文件"""
        all_rows = []
        
        for key, value in data.items():
            if isinstance(value, dict):
                years = value.get('years', [])
                values = value.get('values', [])
                changes = value.get('changes', [])
                
                for year, val, change in zip(years, values, changes):
                    all_rows.append({
                        '指标': value.get('title', key.upper()),
                        '年份': year,
                        '数值': val,
                        '同比变化': change
                    })
        
        if not all_rows:
            filename = f"economic_report_{self._get_timestamp()}.csv"
            filepath = os.path.join(self.output_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("无数据\n")
            return filepath
        
        df = pd.DataFrame(all_rows)
        filename = f"economic_report_{self._get_timestamp()}.csv"
        filepath = os.path.join(self.output_dir, filename)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        return filepath
    
    def _export_excel(self, data: Dict) -> str:
        return self._export_csv_format(data)
    
    def _export_pdf(self, data: Dict) -> str:
        return self._export_csv_format(data)
    
    def _export_json(self, data: Dict) -> str:
        filename = f"economic_report_{self._get_timestamp()}.json"
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump({
                'generated_at': datetime.now().isoformat(),
                'data': data
            }, f, ensure_ascii=False, indent=2)
        
        return filepath


def get_exporter(format: str = 'excel') -> ReportExporter:
    """获取导出器实例
    
    Args:
        format: 导出格式 ('excel', 'pdf', 'csv')
    
    Returns:
        导出器实例
    """
    if format == 'excel':
        return ExcelExporter()
    elif format == 'pdf':
        return PDFExporter()
    else:
        return CSVExporter()


def export_report(data: Dict[str, Any], format: str = 'excel') -> str:
    """导出报告的便捷函数
    
    Args:
        data: 要导出的数据
        format: 导出格式
    
    Returns:
        生成的文件路径
    """
    exporter = get_exporter(format)
    return exporter.export(data, format)


if __name__ == '__main__':
    test_data = {
        'gdp': {
            'title': '国内生产总值',
            'years': [2020, 2021, 2022, 2023, 2024],
            'values': [101.6, 110.4, 121.0, 121.0, 126.1],
            'changes': [2.2, 8.4, 9.5, 3.0, 4.2],
            'latest': '126.1万亿元',
            'change': '+4.2%',
            'trend': '稳中有升'
        },
        'cpi': {
            'title': '居民消费价格指数',
            'years': [2020, 2021, 2022, 2023, 2024],
            'values': [102.5, 100.9, 102.9, 102.0, 101.5],
            'changes': [2.5, 0.9, 2.9, 2.0, 1.5],
            'latest': '101.5',
            'change': '+1.5%',
            'trend': '温和上涨'
        }
    }
    
    print("测试报告导出...")
    
    for fmt in ['csv', 'json']:
        try:
            path = export_report(test_data, fmt)
            print(f"✓ {fmt.upper()} 报告已生成: {path}")
        except Exception as e:
            print(f"✗ {fmt.upper()} 导出失败: {e}")
    
    if HAS_OPENPYXL:
        try:
            path = export_report(test_data, 'excel')
            print(f"✓ Excel 报告已生成: {path}")
        except Exception as e:
            print(f"✗ Excel 导出失败: {e}")
    
    if HAS_REPORTLAB:
        try:
            path = export_report(test_data, 'pdf')
            print(f"✓ PDF 报告已生成: {path}")
        except Exception as e:
            print(f"✗ PDF 导出失败: {e}")
